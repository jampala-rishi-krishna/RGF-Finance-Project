import streamlit as st
import pandas as pd
import re
from datetime import date
from auth import check_login, logout
from database import (init_db, get_lenders, save_lenders, update_lender,
                       get_linkedin_results, get_verified_names, save_linkedin_results)
from scraper import scrape_lenders
from styles import get_css
from linkedin_verifier import verify_batch, get_next_batch, _is_real_company


def _linkedin_html_table(df: pd.DataFrame) -> str:
    """Render LinkedIn results as an HTML table with clickable links."""
    def _parse_members(text: str) -> str:
        if not text or text == "None found":
            return '<span style="color:#6B7068;font-size:0.75rem;">None found</span>'
        parts = text.split(" | ")
        links = []
        for part in parts:
            m = re.search(r'(https?://\S+)', part)
            if m:
                url   = m.group(1).rstrip(")")
                label = part[:m.start()].rstrip(": ").strip() or "View Profile"
                links.append(
                    f'<a href="{url}" target="_blank" '
                    f'style="color:#86000B;text-decoration:none;font-size:0.75rem;">'
                    f'{label}</a>'
                )
            else:
                links.append(f'<span style="font-size:0.75rem;">{part}</span>')
        return '<br>'.join(links)

    TH = "padding:10px 8px;font-size:0.62rem;letter-spacing:2px;text-transform:uppercase;text-align:left;white-space:nowrap;"
    TD = "padding:8px;font-size:0.78rem;vertical-align:top;border-bottom:1px solid #E0DAD0;"

    rows = ""
    for i, (_, row) in enumerate(df.iterrows()):
        bg = "#FFFFFF" if i % 2 == 0 else "#F7F4EF"

        web = str(row.get("Website", "") or "")
        web_html = (f'<a href="{web}" target="_blank" style="color:#86000B;text-decoration:none;">Visit</a>'
                    if web else "—")

        li = str(row.get("LinkedIn Company", "") or "")
        li_html = (f'<a href="{li}" target="_blank" style="color:#86000B;text-decoration:none;">View Page</a>'
                   if li else "—")

        has = str(row.get("Has LinkedIn", "No"))
        has_color = "#1B6337" if has == "Yes" else "#6B7068"

        rows += f"""
        <tr style="background:{bg};">
          <td style="{TD}font-weight:600;">{row.get("Lender Name","")}</td>
          <td style="{TD}">{web_html}</td>
          <td style="{TD}color:{has_color};font-weight:700;">{has}</td>
          <td style="{TD}">{li_html}</td>
          <td style="{TD}">{_parse_members(str(row.get("Member Details","") or ""))}</td>
          <td style="{TD}text-align:center;">{row.get("Members Found",0)}</td>
          <td style="{TD}text-align:center;">{row.get("Fit Score",3)}</td>
          <td style="{TD}color:#6B7068;">{row.get("Type","")}</td>
        </tr>"""

    return f"""
    <div style="overflow-x:auto;border:1px solid #E0DAD0;border-radius:2px;margin-top:8px;">
    <table style="width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;">
      <thead>
        <tr style="background:#1B2419;color:#FFFFFF;">
          <th style="{TH}">Lender Name</th>
          <th style="{TH}">Website</th>
          <th style="{TH}">LinkedIn</th>
          <th style="{TH}">Company Page</th>
          <th style="{TH}">Members / Profiles</th>
          <th style="{TH}text-align:center;">#</th>
          <th style="{TH}text-align:center;">Fit</th>
          <th style="{TH}">Type</th>
        </tr>
      </thead>
      <tbody>{rows}</tbody>
    </table>
    </div>"""


def _to_excel(df: pd.DataFrame) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Lenders"
    cols = list(df.columns)
    header_fill = PatternFill("solid", fgColor="86000B")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_edit_form(lid: int, row):
    with st.expander("EDIT DETAILS", expanded=True):
        e1, e2, e3 = st.columns(3)
        with e1:
            e_contact  = st.text_input("Contact Person",  value=str(row.get("contact_person","") or ""),  key=f"e_cp_{lid}")
            e_position = st.text_input("Position",        value=str(row.get("contact_position","") or ""), key=f"e_pos_{lid}")
            e_email    = st.text_input("Email",           value=str(row.get("contact_email","") or ""),    key=f"e_em_{lid}")
            e_phone    = st.text_input("Phone",           value=str(row.get("contact_phone","") or ""),    key=f"e_ph_{lid}")
        with e2:
            e_rates    = st.text_input("Rates / Fees",    value=str(row.get("rates","") or ""),            key=f"e_rt_{lid}")
            e_ratesrc  = st.text_input("Rate Source URL", value=str(row.get("rate_source_url","") or ""),  key=f"e_rs_{lid}")
            e_loansize = st.text_input("Typical Loan Size", value=str(row.get("typical_loan_size","") or ""), key=f"e_ls_{lid}")
            e_fit      = st.selectbox("Fit Score", [5,3,1],
                                      index=[5,3,1].index(int(row.get("fit_score",3) or 3)),
                                      format_func=lambda x: f"{x} — {'Strong' if x==5 else 'Medium' if x==3 else 'Weak'} Fit",
                                      key=f"e_fit_{lid}")
        with e3:
            e_audit    = st.selectbox("Audited FS Required", ["Unknown","Yes","No"],
                                      index=["Unknown","Yes","No"].index(str(row.get("audited_fs_required","Unknown") or "Unknown")),
                                      key=f"e_aud_{lid}")
            e_profit   = st.selectbox("Profitability Required", ["Unknown","Yes","No"],
                                      index=["Unknown","Yes","No"].index(str(row.get("profitability_required","Unknown") or "Unknown")),
                                      key=f"e_prf_{lid}")
            e_notes    = st.text_area("Notes", value=str(row.get("notes","") or ""), height=100, key=f"e_nt_{lid}")

        if st.button("SAVE CHANGES", key=f"save_{lid}"):
            try:
                update_lender(lid, {
                    "contact_person": e_contact, "contact_position": e_position,
                    "contact_email": e_email, "contact_phone": e_phone,
                    "rates": e_rates, "rate_source_url": e_ratesrc,
                    "typical_loan_size": e_loansize, "fit_score": e_fit,
                    "audited_fs_required": e_audit,
                    "profitability_required": e_profit, "notes": e_notes,
                })
                st.markdown('<div class="alert-success">Lender updated.</div>', unsafe_allow_html=True)
                st.rerun()
            except Exception as ex:
                st.markdown(f'<div class="alert-error">{ex}</div>', unsafe_allow_html=True)


st.set_page_config(
    page_title="RARE Global Food — Lender Intelligence",
    page_icon="R",
    layout="wide",
    initial_sidebar_state="collapsed"
)
st.markdown(get_css(), unsafe_allow_html=True)

# ── Auth ───────────────────────────────────────────────────────────────────
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    from login_page import render_login
    render_login()
    st.stop()

init_db()

if "page" not in st.session_state:
    st.session_state.page = "Lender Database"
page = st.session_state.page

# ── Header ─────────────────────────────────────────────────────────────────
st.markdown("""
<div class="page-header">
    <div class="header-left">
        <div class="header-brand">
            <span class="brand-rare">RARE</span>
            <span class="brand-gf">GLOBAL FOOD</span>
        </div>
        <div class="header-title">Lender Target Database</div>
    </div>
</div>""", unsafe_allow_html=True)

# ── Top Nav Bar ─────────────────────────────────────────────────────────────
st.markdown(
    '<div style="background:#FFFFFF;border-bottom:1px solid #E0DAD0;'
    'height:54px;margin:0 -1rem;margin-bottom:-54px;'
    'position:relative;z-index:0;pointer-events:none;"></div>',
    unsafe_allow_html=True
)
_n1, _n2, _n3, _ngap, _nuser, _nso = st.columns([1.6, 1.2, 1.6, 3.2, 1.8, 0.9])
with _n1:
    if st.button("Lender Database", key="nav_db", use_container_width=True,
                 type="primary" if page == "Lender Database" else "secondary"):
        st.session_state.page = "Lender Database"
        st.rerun()
with _n2:
    if st.button("Run Research", key="nav_rs", use_container_width=True,
                 type="primary" if page == "Run Research" else "secondary"):
        st.session_state.page = "Run Research"
        st.rerun()
with _n3:
    if st.button("Add Manual Entry", key="nav_am", use_container_width=True,
                 type="primary" if page == "Add Manual Entry" else "secondary"):
        st.session_state.page = "Add Manual Entry"
        st.rerun()
with _nuser:
    st.markdown(
        f'<div class="nav-user-label">{st.session_state.get("username","Admin")}</div>',
        unsafe_allow_html=True
    )
with _nso:
    if st.button("Sign Out", key="nav_signout", use_container_width=True, type="secondary"):
        logout()
        st.rerun()

st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

# ── Load all lenders ───────────────────────────────────────────────────────
all_df = get_lenders()

# ── Metrics ────────────────────────────────────────────────────────────────
total      = len(all_df)
strong_fit = len(all_df[all_df["fit_score"] == 5]) if not all_df.empty and "fit_score" in all_df.columns else 0
med_fit    = len(all_df[all_df["fit_score"] == 3]) if not all_df.empty and "fit_score" in all_df.columns else 0
with_contact = len(all_df[all_df["contact_email"].notna() & (all_df["contact_email"] != "")]) if not all_df.empty and "contact_email" in all_df.columns else 0

m1, m2, m3, m4 = st.columns(4)
for col, label, val in [
    (m1, "TOTAL LENDERS",   str(total)),
    (m2, "STRONG FIT (5)",  str(strong_fit)),
    (m3, "MEDIUM FIT (3)",  str(med_fit)),
    (m4, "WITH CONTACT",    str(with_contact)),
]:
    with col:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{val}</div>
        </div>""", unsafe_allow_html=True)

st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════
# PAGE: LENDER DATABASE
# ══════════════════════════════════════════════════════════════════════════
if st.session_state.page == "Lender Database":
    st.markdown('<div class="section-title">LENDER DATABASE</div>', unsafe_allow_html=True)

    # Filters
    fc1, fc2, fc3, fc4 = st.columns([2, 2, 2, 1])
    with fc1:
        search = st.text_input("Search name / website / contact", placeholder="Type to filter...")
    with fc2:
        fit_filter = st.selectbox("Fit Score", ["All", "5 — Strong Fit", "3 — Medium Fit", "1 — Weak Fit"])
    with fc3:
        type_filter = st.selectbox("Lender Type", ["All", "Direct Lender", "Broker", "Marketplace", "Capital Partner"])
    with fc4:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("RESET", use_container_width=True):
            st.rerun()

    display_df = all_df.copy() if not all_df.empty else pd.DataFrame()

    if not display_df.empty:
        if search:
            mask = display_df.apply(lambda r: r.astype(str).str.contains(search, case=False).any(), axis=1)
            display_df = display_df[mask]
        if fit_filter != "All":
            score = int(fit_filter[0])
            display_df = display_df[display_df["fit_score"] == score]
        if type_filter != "All" and "lender_type" in display_df.columns:
            display_df = display_df[display_df["lender_type"] == type_filter]

    DISPLAY_COLS = {
        "legal_name":        "Legal Name",
        "trade_name":        "Trade Name",
        "website":           "Website",
        "lender_type":       "Type",
        "contact_person":    "Contact Person",
        "contact_position":  "Position",
        "contact_email":     "Email",
        "contact_phone":     "Phone",
        "fit_score":         "Fit Score",
        "invoice_finance":   "Invoice Fin.",
        "inventory_finance": "Inventory Fin.",
        "po_finance":        "PO Finance",
        "import_finance":    "Import Fin.",
        "working_capital":   "Working Capital",
        "asset_based":       "Asset Based",
        "source":            "Source",
    }

    if not display_df.empty:
        show_df = display_df[[c for c in DISPLAY_COLS if c in display_df.columns]].rename(columns=DISPLAY_COLS)
        st.markdown(f'<div class="table-count">{len(show_df)} lenders</div>', unsafe_allow_html=True)
        st.dataframe(
            show_df,
            use_container_width=True,
            height=460,
            hide_index=True,
            column_config={
                "Website": st.column_config.LinkColumn("Website", display_text="Visit"),
                "Fit Score": st.column_config.NumberColumn("Fit Score", format="%d"),
            }
        )

        dl1, dl2 = st.columns([1, 4])
        with dl1:
            st.download_button(
                "EXPORT EXCEL",
                data=_to_excel(display_df),
                file_name=f"rare_lenders_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # ── LinkedIn Verification ──────────────────────────────────────────
        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-title">LINKEDIN VERIFICATION</div>', unsafe_allow_html=True)

        # Load persisted results and compute stats
        li_db_df       = get_linkedin_results()
        verified_names = get_verified_names()
        real_total     = int(all_df["legal_name"].apply(lambda n: _is_real_company(str(n))).sum())
        verified_count = len(verified_names)
        remaining      = max(real_total - verified_count, 0)
        next_batch_df  = get_next_batch(all_df, verified_names, batch_size=50)

        # Stats row
        sv1, sv2, sv3 = st.columns(3)
        for col, label, val in [
            (sv1, "REAL COMPANIES",    str(real_total)),
            (sv2, "VERIFIED",          str(verified_count)),
            (sv3, "REMAINING",         str(remaining)),
        ]:
            with col:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-label">{label}</div>
                    <div class="metric-value">{val}</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        st.markdown("""
        <div class="info-box">
            Each click verifies the next 50 unverified lenders and saves results permanently
            to the database — click again to continue with the next 50. Results survive page
            refreshes and are never re-verified (SerpAPI credits are preserved).
        </div>""", unsafe_allow_html=True)

        col_li1, col_li2 = st.columns([1, 4])
        with col_li1:
            verify_btn = st.button(
                "VERIFY NEXT 50",
                use_container_width=True,
                key="verify_linkedin_btn",
                disabled=next_batch_df.empty,
            )
        with col_li2:
            if next_batch_df.empty:
                st.markdown(
                    '<div style="padding-top:10px;font-size:0.75rem;color:#1B6337;">'
                    'All real companies have been verified.</div>',
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f'<div style="padding-top:10px;font-size:0.75rem;color:#6B7068;">'
                    f'Next batch: <strong>{len(next_batch_df)}</strong> unverified lenders. '
                    f'<strong>{verified_count}</strong> already saved — will not be re-queried.</div>',
                    unsafe_allow_html=True
                )

        if verify_btn and not next_batch_df.empty:
            pb_li = st.progress(0, text="Starting LinkedIn verification...")
            try:
                batch_records = verify_batch(next_batch_df, pb_li)
                if batch_records:
                    saved_n = save_linkedin_results(batch_records)
                    pb_li.progress(100, text=f"Done — {saved_n} results saved to database.")
                else:
                    pb_li.progress(100, text="Batch complete — no results returned.")
                st.rerun()
            except Exception as e:
                st.markdown(f'<div class="alert-error">Verification failed: {e}</div>', unsafe_allow_html=True)

        # ── Show all stored results ────────────────────────────────────────
        if not li_db_df.empty:
            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
            has_li_count = int((li_db_df["has_linkedin"] == "Yes").sum()) if "has_linkedin" in li_db_df.columns else 0
            st.markdown(
                f'<div class="table-count">{len(li_db_df)} verified — '
                f'{has_li_count} have LinkedIn</div>',
                unsafe_allow_html=True
            )

            display_li = li_db_df.rename(columns={
                "lender_name":      "Lender Name",
                "website":          "Website",
                "has_linkedin":     "Has LinkedIn",
                "linkedin_company": "LinkedIn Company",
                "members_found":    "Members Found",
                "member_details":   "Member Details",
                "fit_score":        "Fit Score",
                "lender_type":      "Type",
                "verified_at":      "Verified At",
            })

            st.markdown(_linkedin_html_table(display_li), unsafe_allow_html=True)

            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            st.download_button(
                "EXPORT LINKEDIN REPORT",
                data=display_li.to_csv(index=False).encode("utf-8"),
                file_name=f"rare_linkedin_verification_{date.today()}.csv",
                mime="text/csv",
                key="dl_linkedin"
            )

        # ── Lead Detail / Edit ─────────────────────────────────────────────
        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-title">EDIT LENDER RECORD</div>', unsafe_allow_html=True)

        if "id" in display_df.columns:
            name_map = {
                f"{row.get('legal_name','?')} — {row.get('lender_type','')}": row["id"]
                for _, row in display_df.iterrows()
            }
            sel = st.selectbox("Select lender to edit", ["— Select —"] + list(name_map.keys()))

            if sel != "— Select —":
                lid = name_map[sel]
                row = all_df[all_df["id"] == lid].iloc[0]
                _render_edit_form(lid, row)
    else:
        st.markdown('<div class="empty-state">No lenders found. Run Research to populate the database.</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════
# PAGE: RUN RESEARCH
# ══════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "Run Research":
    st.markdown('<div class="section-title">AUTOMATED LENDER RESEARCH</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="info-box">
        This tool searches BSP Directory, SEC Registry, and Google to find active lenders in the Philippines.
        Results are automatically saved to the database. Run each source independently.
    </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    # Source selection
    src1, src2, src3 = st.columns(3)
    with src1:
        st.markdown("""
        <div class="source-card">
            <div class="source-title">BSP DIRECTORY</div>
            <div class="source-desc">Bank of the Philippines Islands licensed banks, quasi-banks, and financing companies. Most reliable source.</div>
        </div>""", unsafe_allow_html=True)
        run_bsp = st.button("RUN BSP SCRAPE", use_container_width=True, key="bsp")

    with src2:
        st.markdown("""
        <div class="source-card">
            <div class="source-title">SEC REGISTRY</div>
            <div class="source-desc">Securities and Exchange Commission registered lending and financing companies in the Philippines.</div>
        </div>""", unsafe_allow_html=True)
        run_sec = st.button("RUN SEC SCRAPE", use_container_width=True, key="sec")

    with src3:
        st.markdown("""
        <div class="source-card">
            <div class="source-title">GOOGLE SEARCH</div>
            <div class="source-desc">Search Google for trade finance, invoice finance, and inventory finance providers in Philippines.</div>
        </div>""", unsafe_allow_html=True)
        run_google = st.button("RUN GOOGLE SEARCH", use_container_width=True, key="google")

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    if run_bsp:
        pb = st.progress(0, text="Scraping BSP Directory...")
        try:
            df = scrape_lenders("bsp", pb)
            if df is not None and not df.empty:
                saved, dupes = save_lenders(df)
                pb.progress(100, text="Done.")
                st.markdown(f'<div class="alert-success">BSP scrape complete — {saved} new lenders saved. {dupes} duplicates skipped.</div>', unsafe_allow_html=True)
                st.rerun()
            else:
                st.markdown('<div class="alert-warn">No results from BSP. Try again later.</div>', unsafe_allow_html=True)
        except Exception as e:
            st.markdown(f'<div class="alert-error">BSP scrape failed: {e}</div>', unsafe_allow_html=True)

    if run_sec:
        pb = st.progress(0, text="Scraping SEC Registry...")
        try:
            df = scrape_lenders("sec", pb)
            if df is not None and not df.empty:
                saved, dupes = save_lenders(df)
                pb.progress(100, text="Done.")
                st.markdown(f'<div class="alert-success">SEC scrape complete — {saved} new lenders saved. {dupes} duplicates skipped.</div>', unsafe_allow_html=True)
                st.rerun()
            else:
                st.markdown('<div class="alert-warn">No results from SEC. Try again later.</div>', unsafe_allow_html=True)
        except Exception as e:
            st.markdown(f'<div class="alert-error">SEC scrape failed: {e}</div>', unsafe_allow_html=True)

    if run_google:
        pb = st.progress(0, text="Running Google searches...")
        try:
            df = scrape_lenders("google", pb)
            if df is not None and not df.empty:
                saved, dupes = save_lenders(df)
                pb.progress(100, text="Done.")
                st.markdown(f'<div class="alert-success">Google search complete — {saved} new lenders saved. {dupes} duplicates skipped.</div>', unsafe_allow_html=True)
                st.rerun()
            else:
                st.markdown('<div class="alert-warn">No results from Google search.</div>', unsafe_allow_html=True)
        except Exception as e:
            st.markdown(f'<div class="alert-error">Google search failed: {e}</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════
# PAGE: ADD MANUAL ENTRY
# ══════════════════════════════════════════════════════════════════════════
elif st.session_state.page == "Add Manual Entry":
    st.markdown('<div class="section-title">ADD LENDER MANUALLY</div>', unsafe_allow_html=True)

    with st.container():
        st.markdown('<div class="control-panel">', unsafe_allow_html=True)

        a1, a2, a3 = st.columns(3)
        with a1:
            m_legal    = st.text_input("Legal Name *", key="m_legal")
            m_trade    = st.text_input("Trade Name", key="m_trade")
            m_website  = st.text_input("Website", key="m_website")
            m_linkedin = st.text_input("LinkedIn Company Page", key="m_linkedin")
            m_status   = st.selectbox("Status", ["Active", "Dormant", "Unknown"], key="m_status")
        with a2:
            m_contact  = st.text_input("Contact Person", key="m_contact")
            m_position = st.text_input("Position / Title", key="m_position")
            m_email    = st.text_input("Email", key="m_email")
            m_phone    = st.text_input("Phone", key="m_phone")
            m_li_prof  = st.text_input("LinkedIn Profile URL", key="m_li_prof")
        with a3:
            m_type     = st.selectbox("Lender Type", ["Direct Lender", "Broker", "Marketplace", "Capital Partner"], key="m_type")
            m_fit      = st.selectbox("Rare Fit Score", [5, 3, 1], format_func=lambda x: f"{x} — {'Strong' if x==5 else 'Medium' if x==3 else 'Weak'} Fit", key="m_fit")
            m_rates    = st.text_input("Rates / Fees (or 'Not Disclosed')", key="m_rates")
            m_rate_src = st.text_input("Rate Source URL", key="m_rate_src")
            m_notes    = st.text_area("Notes", height=80, key="m_notes")

        st.markdown("**Products Offered**")
        p1, p2, p3, p4, p5, p6 = st.columns(6)
        m_inv_fin  = p1.checkbox("Invoice Finance",    key="m_inv_fin")
        m_inv_ory  = p2.checkbox("Inventory Finance",  key="m_inv_ory")
        m_po       = p3.checkbox("PO Finance",         key="m_po")
        m_import   = p4.checkbox("Import Finance",     key="m_import")
        m_wc       = p5.checkbox("Working Capital",    key="m_wc")
        m_asset    = p6.checkbox("Asset Based",        key="m_asset")

        st.markdown("**Borrower Requirements**")
        r1, r2, r3, r4 = st.columns(4)
        m_yrs      = r1.number_input("Min Years in Business", min_value=0, max_value=50, value=0, key="m_yrs")
        m_audit    = r2.selectbox("Audited FS Required", ["Unknown", "Yes", "No"], key="m_audit")
        m_profit   = r3.selectbox("Profitability Required", ["Unknown", "Yes", "No"], key="m_profit")
        m_loansize = r4.text_input("Typical Loan Size", key="m_loansize")

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        if st.button("SAVE LENDER", use_container_width=False, key="save_manual"):
            if not m_legal:
                st.markdown('<div class="alert-error">Legal Name is required.</div>', unsafe_allow_html=True)
            else:
                record = pd.DataFrame([{
                    "legal_name": m_legal, "trade_name": m_trade,
                    "website": m_website, "linkedin_company": m_linkedin,
                    "active_status": m_status,
                    "contact_person": m_contact, "contact_position": m_position,
                    "contact_email": m_email, "contact_phone": m_phone,
                    "linkedin_profile": m_li_prof,
                    "lender_type": m_type, "fit_score": m_fit,
                    "invoice_finance": m_inv_fin, "inventory_finance": m_inv_ory,
                    "po_finance": m_po, "import_finance": m_import,
                    "working_capital": m_wc, "asset_based": m_asset,
                    "min_years_business": m_yrs, "audited_fs_required": m_audit,
                    "profitability_required": m_profit, "typical_loan_size": m_loansize,
                    "rates": m_rates, "rate_source_url": m_rate_src,
                    "notes": m_notes, "source": "Manual Entry",
                }])
                saved, dupes = save_lenders(record)
                if saved:
                    st.markdown('<div class="alert-success">Lender saved successfully.</div>', unsafe_allow_html=True)
                    st.rerun()
                else:
                    st.markdown('<div class="alert-warn">Lender already exists in the database.</div>', unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

